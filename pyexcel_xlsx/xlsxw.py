"""
    pyexcel_xlsx.xlsxw
    ~~~~~~~~~~~~~~~~~~~

    Write xlsx file format using openpyxl

    :copyright: (c) 2015-2017 by Onni Software Ltd & its contributors
    :license: New BSD License
"""
import openpyxl

from pyexcel_io.book import BookWriter
from pyexcel_io.sheet import SheetWriter


class XLSXSheetWriter(SheetWriter):
    """
    xls, xlsx and xlsm sheet writer
    """
    def set_sheet_name(self, name):
        self._native_sheet.title = name
        self.current_row = 1

    def write_row(self, array):
        """
        write a row into the file
        """
        self._native_sheet.append(array)


class XLSXWriter(BookWriter):
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self):
        BookWriter.__init__(self)
        self.current_sheet = 0
        self._native_book = None

    def open(self, file_name, **keywords):
        BookWriter.open(self, file_name, **keywords)
        self._native_book = openpyxl.Workbook(write_only=True)

    def create_sheet(self, name):
        return XLSXSheetWriter(self._native_book,
                               self._native_book.create_sheet(), name)

    def close(self):
        """
        This call actually save the file
        """
        self._native_book.save(filename=self._file_alike_object)
        self._native_book = None
