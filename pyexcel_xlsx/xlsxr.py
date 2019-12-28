"""
    pyexcel_xlsx.xlsxr
    ~~~~~~~~~~~~~~~~~~~

    Read xlsx file format using openpyxl

    :copyright: (c) 2015-2019 by Onni Software Ltd & its contributors
    :license: New BSD License
"""
import openpyxl

from pyexcel_io.book import BookReader
from pyexcel_io.sheet import SheetReader
from pyexcel_io._compact import OrderedDict, irange


class FastSheet(SheetReader):
    """
    Iterate through rows
    """

    @property
    def name(self):
        """sheet name"""
        return self._native_sheet.title

    def row_iterator(self):
        """
        openpyxl row iterator

        http://openpyxl.readthedocs.io/en/default/optimized.html
        """
        for row in self._native_sheet.rows:
            yield row

    def column_iterator(self, row):
        """
        a generator for the values in a row
        """
        for cell in row:
            yield cell.value


class MergedCell(object):
    def __init__(self, cell_ranges):
        self.__cl, self.__rl, self.__ch, self.__rh = cell_ranges.bounds
        self.value = None

    def register_cells(self, registry):
        for rowx in irange(self.__rl, self.__rh + 1):
            for colx in irange(self.__cl, self.__ch + 1):
                key = "%s-%s" % (rowx, colx)
                registry[key] = self

    def bottom_row(self):
        return self.__rh

    def right_column(self):
        return self.__ch


class SlowSheet(FastSheet):
    """
    This sheet will be slower because it does not use readonly sheet
    """

    def __init__(self, sheet, **keywords):
        SheetReader.__init__(self, sheet, **keywords)
        self.__merged_cells = {}
        self.max_row = 0
        self.max_column = 0
        self.__sheet_max_row = sheet.max_row
        self.__sheet_max_column = sheet.max_column
        for ranges in sheet.merged_cells.ranges[:]:
            merged_cells = MergedCell(ranges)
            merged_cells.register_cells(self.__merged_cells)
            if self.max_row < merged_cells.bottom_row():
                self.max_row = merged_cells.bottom_row()
            if self.max_column < merged_cells.right_column():
                self.max_column = merged_cells.right_column()

    def row_iterator(self):
        """
        skip hidden rows
        """
        for row_index, row in enumerate(self._native_sheet.rows, 1):
            if self._native_sheet.row_dimensions[row_index].hidden is False:
                yield (row, row_index)
        if self.max_row > self.__sheet_max_row:
            for i in range(self.__sheet_max_row, self.max_row):
                data = [None] * self.__sheet_max_column
                yield (data, i + 1)

    def column_iterator(self, row_struct):
        """
        skip hidden columns
        """
        row, row_index = row_struct
        for column_index, cell in enumerate(row, 1):
            letter = openpyxl.utils.get_column_letter(column_index)
            if self._native_sheet.column_dimensions[letter].hidden is False:
                if cell:
                    value = cell.value
                else:
                    value = ""
                if value is None:
                    value = ""
                value = self._merged_cells(row_index, column_index, value)
                yield value
        if self.max_column > self.__sheet_max_column:
            for i in range(self.__sheet_max_column, self.max_column):
                value = self._merged_cells(row_index, i + 1, "")
                yield value

    def _merged_cells(self, row, column, value):
        ret = value
        if self.__merged_cells:
            merged_cell = self.__merged_cells.get("%s-%s" % (row, column))
            if merged_cell:
                if merged_cell.value:
                    ret = merged_cell.value
                else:
                    merged_cell.value = value
        return ret


class XLSXBook(BookReader):
    """
    Open xlsx as read only mode
    """

    def open(
        self,
        file_name,
        skip_hidden_sheets=True,
        detect_merged_cells=False,
        skip_hidden_row_and_column=True,
        **keywords
    ):

        BookReader.open(self, file_name, **keywords)
        self.skip_hidden_sheets = skip_hidden_sheets
        self.skip_hidden_row_and_column = skip_hidden_row_and_column
        self.detect_merged_cells = detect_merged_cells
        self._load_the_excel_file(file_name)

    def open_stream(
        self,
        file_stream,
        skip_hidden_sheets=True,
        detect_merged_cells=False,
        skip_hidden_row_and_column=True,
        **keywords
    ):
        BookReader.open_stream(self, file_stream, **keywords)
        self.skip_hidden_sheets = skip_hidden_sheets
        self.skip_hidden_row_and_column = skip_hidden_row_and_column
        self.detect_merged_cells = detect_merged_cells
        self._load_the_excel_file(file_stream)

    def read_sheet_by_name(self, sheet_name):
        sheet = self._native_book[sheet_name]
        if sheet is None:
            raise ValueError("%s cannot be found" % sheet_name)
        else:
            return self.read_sheet(sheet)

    def read_sheet_by_index(self, sheet_index):
        names = self._native_book.sheetnames
        length = len(names)
        if sheet_index < length:
            return self.read_sheet_by_name(names[sheet_index])
        else:
            raise IndexError(
                "Index %d of out bound %d" % (sheet_index, length)
            )

    def read_all(self):
        result = OrderedDict()
        for sheet in self._native_book:
            if self.skip_hidden_sheets and sheet.sheet_state == "hidden":
                continue
            data_dict = self.read_sheet(sheet)
            result.update(data_dict)
        return result

    def read_sheet(self, native_sheet):
        if self.skip_hidden_row_and_column or self.detect_merged_cells:
            sheet = SlowSheet(native_sheet, **self._keywords)
        else:
            sheet = FastSheet(native_sheet, **self._keywords)
        return {sheet.name: sheet.to_array()}

    def close(self):
        self._native_book.close()
        self._native_book = None

    def _load_the_excel_file(self, file_alike_object):
        read_only_flag = True
        if self.skip_hidden_row_and_column:
            read_only_flag = False
        data_only_flag = True
        if self.detect_merged_cells:
            data_only_flag = False
        self._native_book = openpyxl.load_workbook(
            filename=file_alike_object,
            data_only=data_only_flag,
            read_only=read_only_flag,
        )
