"""
    pyexcel.ext.xlbook
    ~~~~~~~~~~~~~~~~~~~

    The lower level xls/xlsx/xlsm file format handler using xlrd/xlwt

    :copyright: (c) 2014 by C. W.
    :license: GPL v3
"""
import sys
import datetime
import openpyxl
if sys.version_info[0] < 3:
    from StringIO import StringIO
else:
    from io import BytesIO as StringIO
if sys.version_info[0] == 2 and sys.version_info[1] < 7:
    from ordereddict import OrderedDict
else:
    from collections import OrderedDict


COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
COLUMN_LENGTH = 26

def get_columns(index):
    if index < COLUMN_LENGTH:
        return COLUMNS[index]
    else:
        return get_columns(index/COLUMN_LENGTH) + COLUMNS[index%COLUMN_LENGTH]
    

class XLSXSheet:
    """
    xls sheet

    Currently only support first sheet in the file
    """
    def __init__(self, sheet):
        self.worksheet = sheet

    def number_of_rows(self):
        """
        Number of rows in the xls sheet
        """
        return self.worksheet.get_highest_row()

    def number_of_columns(self):
        """
        Number of columns in the xls sheet
        """
        return self.worksheet.get_highest_column()

    def cell_value(self, row, column):
        """
        Random access to the xls cells
        """
        actual_row = row + 1
        return self.worksheet.cell("%s%d" % (get_columns(column), actual_row)).value


def to_array(sheet):
    array = []
    for r in range(0, sheet.number_of_rows()):
        row = []
        for c in range(0, sheet.number_of_columns()):
            row.append(sheet.cell_value(r, c))
        array.append(row)
    return array


class XLSXBook:
    """
    XLSBook reader

    It reads xls, xlsm, xlsx work book
    """

    def __init__(self, filename, file_content=None, **keywords):
        if file_content:
            self.workbook = openpyxl.load_workbook(filename=StringIO(file_content))
        else:
            self.workbook = openpyxl.load_workbook(filename)
        self.mysheets = OrderedDict()
        for sheet in self.workbook:
            data = to_array(XLSXSheet(sheet))
            self.mysheets[sheet.title] = data

    def sheets(self):
        """Get sheets in a dictionary"""
        return self.mysheets


class XLSXSheetWriter:
    """
    xls, xlsx and xlsm sheet writer
    """
    def __init__(self, sheet, name):
        if name:
            sheet_name = name
        else:
            sheet_name = "pyexcel_sheet1"
        self.ws = sheet
        self.ws.title = sheet_name
        self.current_row = 1

    def set_size(self, size):
        pass

    def write_row(self, array):
        """
        write a row into the file
        """
        for i in range(0, len(array)):
            value = array[i]
            style = None
            tmp_array = []
            self.ws.cell("%s%d" % (get_columns(i), self.current_row)).value = value
        self.current_row += 1

    def write_array(self, table):
        for r in table:
            self.write_row(r)

    def close(self):
        """
        This call actually save the file
        """
        pass


class XLSXWriter:
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self, file):
        self.file = file
        self.wb = openpyxl.Workbook()
        self.current_sheet = 0

    def create_sheet(self, name):
        if self.current_sheet == 0:
            self.current_sheet = 1
            return XLSXSheetWriter(self.wb.active, name)
        else:
            
            return XLSXSheetWriter(self.wb.create_sheet(), name)
            

    def write(self, sheet_dicts):
        """Write a dictionary to a multi-sheet file

        Requirements for the dictionary is: key is the sheet name,
        its value must be two dimensional array
        """
        keys = sheet_dicts.keys()
        for name in keys:
            sheet = self.create_sheet(name)
            sheet.write_array(sheet_dicts[name])

    def close(self):
        """
        This call actually save the file
        """
        self.wb.save(filename=self.file)
