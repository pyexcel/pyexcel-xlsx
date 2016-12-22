"""
    pyexcel_xlsx
    ~~~~~~~~~~~~~~~~~~~

    The lower level xlsx file format handler using openpyxl

    :copyright: (c) 2015-2016 by Onni Software Ltd & its contributors
    :license: New BSD License
"""
import sys
from pyexcel_io.book import BookReader, BookWriter
from pyexcel_io.sheet import SheetReader, SheetWriter
import openpyxl

PY27_BELOW = sys.version_info[0] == 2 and sys.version_info[1] < 7
if PY27_BELOW:
    from ordereddict import OrderedDict
else:
    from collections import OrderedDict


COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
COLUMN_LENGTH = 26


def get_columns(index):
    """
    Convert column index to column name
    """
    if index < COLUMN_LENGTH:
        return COLUMNS[index]
    else:
        return (get_columns(int(index // COLUMN_LENGTH) - 1) +
                COLUMNS[index % COLUMN_LENGTH])


class XLSXSheet(SheetReader):
    """
    xls sheet
    """
    @property
    def name(self):
        """sheet name"""
        return self._native_sheet.title

    def row_iterator(self):
        """
        Number of rows in the xls sheet
        """
        return self._native_sheet.rows

    def column_iterator(self, row):
        """
        Number of columns in the xls sheet
        """
        for cell in row:
            yield cell.value


class XLSXBook(BookReader):
    """
    XLSBook reader

    It reads xls, xlsm, xlsx work book
    """
    def open(self, file_name, **keywords):
        BookReader.open(self, file_name, **keywords)
        self._get_params()
        self._load_the_excel_file(file_name)

    def open_stream(self, file_stream, **keywords):
        BookReader.open_stream(self, file_stream, **keywords)
        self._get_params()
        self._load_the_excel_file(file_stream)

    def read_sheet_by_name(self, sheet_name):
        sheet = self._native_book.get_sheet_by_name(sheet_name)
        if sheet is None:
            raise ValueError("%s cannot be found" % sheet_name)
        else:
            return self.read_sheet(sheet)

    def read_sheet_by_index(self, sheet_index):
        names = self._native_book.sheetnames
        length = len(names)
        if sheet_index < length:
            return self.read_sheet_by_name(names[sheet_index])
        else:
            raise IndexError("Index %d of out bound %d" % (
                sheet_index,
                length))

    def read_all(self):
        result = OrderedDict()
        for sheet in self._native_book:
            if self.skip_hidden_sheets and sheet.sheet_state == 'hidden':
                continue
            data_dict = self.read_sheet(sheet)
            result.update(data_dict)
        return result

    def read_sheet(self, native_sheet):
        sheet = XLSXSheet(native_sheet, **self._keywords)
        return {sheet.name: sheet.to_array()}

    def _load_the_excel_file(self, file_alike_object):
        self._native_book = openpyxl.load_workbook(
            filename=file_alike_object, data_only=True, read_only=True)

    def _get_params(self):
        self.skip_hidden_sheets = self._keywords.get(
            'skip_hidden_sheets', True)


class XLSXSheetWriter(SheetWriter):
    """
    xls, xlsx and xlsm sheet writer
    """
    def set_sheet_name(self, name):
        self._native_sheet.title = name
        self.current_row = 1

    def write_row(self, array):
        """
        write a row into the file
        """
        self._native_sheet.append(array)


class XLSXWriter(BookWriter):
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self):
        BookWriter.__init__(self)
        self.current_sheet = 0
        self._native_book = None

    def open(self, file_name, **keywords):
        BookWriter.open(self, file_name, **keywords)
        self._native_book = openpyxl.Workbook(write_only=True)

    def create_sheet(self, name):
        return XLSXSheetWriter(self._native_book,
                               self._native_book.create_sheet(), name)

    def close(self):
        """
        This call actually save the file
        """
        self._native_book.save(filename=self._file_alike_object)


_XLSX_MIME = (
    "application/" +
    "vnd.openxmlformats-officedocument.spreadsheetml.sheet")

_xlsx_registry = {
    "file_type": "xlsx",
    "reader": XLSXBook,
    "writer": XLSXWriter,
    "stream_type": "binary",
    "mime_type": _XLSX_MIME,
    "library": "pyexcel-xlsx"
}

_xlsm_registry = {
    "file_type": "xlsm",
    "reader": XLSXBook,
    "writer": XLSXWriter,
    "stream_type": "binary",
    "mime_type": "application/vnd.ms-excel.sheet.macroenabled.12",
    "library": "pyexcel-xlsx"
}

exports = (_xlsx_registry, _xlsm_registry)
