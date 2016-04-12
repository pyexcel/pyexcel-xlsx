"""
    pyexcel_xlsx
    ~~~~~~~~~~~~~~~~~~~

    The lower level xlsx file format handler using openpyxl

    :copyright: (c) 2015-2016 by Onni Software Ltd & its contributors
    :license: New BSD License
"""
import sys
import openpyxl

from pyexcel_io.io import get_data as read_data, isstream, store_data as write_data
from pyexcel_io.book import BookReader, BookWriter
from pyexcel_io.sheet import SheetReader, SheetWriter
from pyexcel_io.manager import RWManager

PY2 = sys.version_info[0] == 2
if PY2 and sys.version_info[1] < 7:
    from ordereddict import OrderedDict
else:
    from collections import OrderedDict    


COLUMNS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
COLUMN_LENGTH = 26


def get_columns(index):
    if index < COLUMN_LENGTH:
        return COLUMNS[index]
    else:
        return (get_columns(int(index // COLUMN_LENGTH) - 1) + COLUMNS[index % COLUMN_LENGTH])


class XLSXSheet(SheetReader):
    """
    xls sheet
    """
    @property
    def name(self):
        return self.native_sheet.title

    def number_of_rows(self):
        """
        Number of rows in the xls sheet
        """
        return self.native_sheet.max_row

    def number_of_columns(self):
        """
        Number of columns in the xls sheet
        """
        return self.native_sheet.max_column

    def cell_value(self, row, column):
        """
        Random access to the xls cells
        """
        actual_row = row + 1
        cell_location = "%s%d" % (get_columns(column), actual_row)
        return self.native_sheet.cell(cell_location).value

    def to_array(self):
        for r in range(0, self.number_of_rows()):
            row = []
            tmp_row = []
            for c in range(0, self.number_of_columns()):
                cell_value = self.cell_value(r, c)
                tmp_row.append(cell_value)
                if cell_value is not None and cell_value != '':
                    row += tmp_row
                    tmp_row = []
            yield row


class XLSXBook(BookReader):
    """
    XLSBook reader

    It reads xls, xlsm, xlsx work book
    """
    def __init__(self):
        BookReader.__init__(self, 'xlsx')
        self.book = None

    def open(self, file_name, **keywords):
        BookReader.open(self, file_name, **keywords)
        self._load_from_file()

    def open_stream(self, file_stream, **keywords):
        BookReader.open_stream(self, file_stream, **keywords)
        self._load_from_memory()

    def read_sheet_by_name(self, sheet_name):
        sheet = self.book.get_sheet_by_name(sheet_name)
        if sheet is None:
            raise ValueError("%s cannot be found" % sheet_name)
        else:
            sheet = XLSXSheet(sheet)
            return {sheet_name: sheet.to_array()}

    def read_sheet_by_index(self, sheet_index):
        names = self.book.sheetnames
        length = len(names)
        if sheet_index < length:
            return self.read_sheet_by_name(names[sheet_index])
        else:
            raise IndexError("Index %d of out bound %d" %(
                sheet_index,
                length))

    def read_all(self):
        result = OrderedDict()
        for sheet in self.book:
            sheet = XLSXSheet(sheet)
            result[sheet.name] = sheet.to_array()
        return result
        
    def _load_from_memory(self):
        self.book =  openpyxl.load_workbook(filename=self.file_stream,
                                            data_only=True)

    def _load_from_file(self):
        self.book = openpyxl.load_workbook(filename=self.file_name,
                                           data_only=True)


class XLSXSheetWriter(SheetWriter):
    """
    xls, xlsx and xlsm sheet writer
    """
    def set_sheet_name(self, name):
        self.native_sheet.title = name
        self.current_row = 1

    def write_row(self, array):
        """
        write a row into the file
        """
        for i in range(0, len(array)):
            cell_location = "%s%d" % (get_columns(i), self.current_row)
            self.native_sheet.cell(cell_location).value = array[i]
        self.current_row += 1


class XLSXWriter(BookWriter):
    """
    xls, xlsx and xlsm writer
    """
    def __init__(self):
        BookWriter.__init__(self, 'xlsx')
        self.current_sheet = 0
        self.native_book = None

    def open(self, file_name, **keywords):
        BookWriter.open(self, file_name, **keywords)
        self.native_book = openpyxl.Workbook()

    def create_sheet(self, name):
        if self.current_sheet == 0:
            self.current_sheet = 1
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.active, name)
        else:
            return XLSXSheetWriter(self.native_book,
                                   self.native_book.create_sheet(), name)

    def close(self):
        """
        This call actually save the file
        """
        self.native_book.save(filename=self.file_alike_object)


def save_data(afile, data, file_type=None, **keywords):
    if isstream(afile) and file_type is None:
        file_type='xlsx'
    write_data(afile, data, file_type=file_type, **keywords)


def get_data(afile, file_type=None, **keywords):
    if isstream(afile) and file_type is None:
        file_type='xlsx'
    return read_data(afile, file_type=file_type, **keywords)



RWManager.register_readers(
    {
        "xlsm": XLSXBook,
        "xlsx": XLSXBook
    })
RWManager.register_writers(
    {
        "xlsm": XLSXWriter,
        "xlsx": XLSXWriter
    })
RWManager.register_file_type_as_binary_stream('xlsm')
RWManager.register_file_type_as_binary_stream('xlsx')
